"""Data import scripts for AMR-Guard structured documents."""

import pandas as pd
from pathlib import Path
from .database import (
    get_connection, init_database, execute_many,
    DOCS_DIR, DB_PATH
)


def safe_float(value):
    """Convert value to float; return None if the value is NaN or non-numeric."""
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_int(value):
    """Convert value to int via float; return None if the value is NaN or non-numeric."""
    if pd.isna(value):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def safe_str(value) -> str:
    """Convert value to string; return empty string for None or NaN."""
    if value is None or pd.isna(value):
        return ''
    return str(value)


def classify_severity(description: str) -> str:
    """
    Classify drug interaction severity from the interaction description text.

    Returns 'major', 'moderate', or 'minor' based on keyword presence.
    Major keywords take precedence over moderate.
    """
    if not description:
        return "unknown"

    desc_lower = description.lower()

    major_keywords = [
        "cardiotoxic", "nephrotoxic", "hepatotoxic", "neurotoxic",
        "fatal", "death", "severe", "contraindicated", "arrhythmia",
        "qt prolongation", "seizure", "bleeding", "hemorrhage",
        "serotonin syndrome", "neuroleptic malignant",
    ]
    moderate_keywords = [
        "increase", "decrease", "reduce", "enhance", "inhibit",
        "metabolism", "concentration", "absorption", "excretion",
        "therapeutic effect", "adverse effect", "toxicity",
    ]

    if any(kw in desc_lower for kw in major_keywords):
        return "major"
    if any(kw in desc_lower for kw in moderate_keywords):
        return "moderate"
    return "minor"


def import_eml_antibiotics() -> int:
    """Import WHO EML antibiotic classification data from the three AWaRe Excel files."""
    print("Importing EML antibiotic data...")

    eml_files = {
        "ACCESS": DOCS_DIR / "antibiotic_guidelines" / "EML export-ACCESS group.xlsx",
        "RESERVE": DOCS_DIR / "antibiotic_guidelines" / "EML export-RESERVE group.xlsx",
        "WATCH": DOCS_DIR / "antibiotic_guidelines" / "EML export-WATCH group.xlsx",
    }

    records = []
    for category, filepath in eml_files.items():
        if not filepath.exists():
            print(f"  Warning: {filepath} not found, skipping...")
            continue

        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active

            headers = [
                str(cell.value).strip().lower().replace(' ', '_') if cell.value else f'col_{i}'
                for i, cell in enumerate(ws[1])
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                medicine = str(row_dict.get('medicine_name', row_dict.get('medicine', '')))
                if not medicine or medicine in ('None', 'nan'):
                    continue

                records.append((
                    medicine,
                    category,
                    safe_str(row_dict.get('eml_section', '')),
                    safe_str(row_dict.get('formulations', '')),
                    safe_str(row_dict.get('indication', '')),
                    safe_str(row_dict.get('atc_codes', row_dict.get('atc_code', ''))),
                    safe_str(row_dict.get('combined_with', '')),
                    safe_str(row_dict.get('status', '')),
                ))

            wb.close()
            print(f"  Loaded {sum(1 for r in records if r[1] == category)} from {category}")

        except Exception as e:
            print(f"  Warning: Error reading {filepath}: {e}")
            continue

    if records:
        execute_many(
            """INSERT INTO eml_antibiotics
               (medicine_name, who_category, eml_section, formulations,
                indication, atc_codes, combined_with, status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            records,
        )
        print(f"  Imported {len(records)} EML antibiotic records total")

    return len(records)


def import_atlas_susceptibility() -> int:
    """Import ATLAS antimicrobial susceptibility data."""
    print("Importing ATLAS susceptibility data...")

    filepath = DOCS_DIR / "pathogen_resistance" / "ATLAS Susceptibility Data Export.xlsx"

    if not filepath.exists():
        print(f"  Warning: {filepath} not found, skipping...")
        return 0

    df_raw = pd.read_excel(filepath, sheet_name="Percent", header=None)

    # Title row contains "Percentage Susceptibility from <Country>"
    region = "Unknown"
    for _, row in df_raw.head(5).iterrows():
        cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        if "from" in cell.lower():
            parts = cell.split("from")
            if len(parts) > 1:
                region = parts[1].strip()
            break

    # Locate the actual header row by finding "Antibacterial"
    header_row = 4
    for idx, row in df_raw.head(10).iterrows():
        if any('Antibacterial' in str(v) for v in row.values if pd.notna(v)):
            header_row = idx
            break

    df = pd.read_excel(filepath, sheet_name="Percent", header=header_row)
    df.columns = [str(col).strip().lower().replace(' ', '_').replace('.', '') for col in df.columns]

    records = []
    for _, row in df.iterrows():
        antibiotic = str(row.get('antibacterial', ''))
        if not antibiotic or antibiotic == 'nan' or 'omitted' in antibiotic.lower():
            continue
        if 'in vitro' in antibiotic.lower() or 'table cells' in antibiotic.lower():
            continue

        n_int = safe_int(row.get('n'))
        s_float = safe_float(row.get('susc', row.get('susceptible')))

        if n_int is not None and s_float is not None:
            records.append((
                "General",
                "",
                antibiotic,
                s_float,
                safe_float(row.get('int', row.get('intermediate'))),
                safe_float(row.get('res', row.get('resistant'))),
                n_int,
                2024,
                region,
                "ATLAS",
            ))

    if records:
        execute_many(
            """INSERT INTO atlas_susceptibility
               (species, family, antibiotic, percent_susceptible,
                percent_intermediate, percent_resistant, total_isolates,
                year, region, source)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            records,
        )
        print(f"  Imported {len(records)} ATLAS susceptibility records from {region}")

    return len(records)


def import_mic_breakpoints() -> int:
    """Import EUCAST MIC breakpoint tables from the Excel file."""
    print("Importing MIC breakpoint data...")

    filepath = DOCS_DIR / "mic_breakpoints" / "v_16.0__BreakpointTables.xlsx"
    if not filepath.exists():
        print(f"  Warning: {filepath} not found, skipping...")
        return 0

    xl = pd.ExcelFile(filepath)
    # These sheets contain metadata/guidance, not pathogen-specific breakpoints
    skip_sheets = {'Content', 'Changes', 'Notes', 'Guidance', 'Dosages',
                   'Technical uncertainty', 'PK PD breakpoints', 'PK PD cutoffs'}

    records = []
    for sheet_name in xl.sheet_names:
        if sheet_name in skip_sheets:
            continue
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            for _, row in df.iterrows():
                row_values = [str(v).strip() for v in row.values if pd.notna(v)]
                if len(row_values) < 2:
                    continue

                potential_antibiotic = row_values[0]
                if any(kw in potential_antibiotic.lower() for kw in
                       ['antibiotic', 'agent', 'note', 'disk', 'mic', 'breakpoint']):
                    continue

                # Extract numeric MIC values; strip inequality signs
                mic_values = []
                for v in row_values[1:]:
                    try:
                        mic_values.append(float(v.replace('≤', '').replace('>', '').replace('<', '').strip()))
                    except (ValueError, AttributeError):
                        pass

                if len(mic_values) >= 2 and len(potential_antibiotic) > 2:
                    records.append((
                        sheet_name,          # pathogen_group
                        potential_antibiotic,
                        None,                # route
                        mic_values[0],       # S breakpoint
                        mic_values[1],       # R breakpoint
                        None, None, None,    # disk S, disk R, notes
                        "16.0",
                    ))
        except Exception as e:
            print(f"  Warning: Could not parse sheet '{sheet_name}': {e}")
            continue

    if records:
        execute_many(
            """INSERT INTO mic_breakpoints
               (pathogen_group, antibiotic, route, mic_susceptible, mic_resistant,
                disk_susceptible, disk_resistant, notes, eucast_version)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            records,
        )
        print(f"  Imported {len(records)} MIC breakpoint records")

    return len(records)


KAGGLE_DATASET = "mghobashy/drug-drug-interactions"
KAGGLE_INPUT_DIR = Path("/kaggle/input/drug-drug-interactions")
INTERACTIONS_CSV = DOCS_DIR / "drug_safety" / "db_drug_interactions.csv"


def _resolve_interactions_csv() -> Path | None:
    """
    Find the drug interactions CSV file.

    Checks in order:
    1. docs/drug_safety/db_drug_interactions.csv (local)
    2. /kaggle/input/drug-drug-interactions/ (Kaggle notebook with dataset attached)
    3. Kaggle API download (requires ~/.kaggle/kaggle.json)
    """
    if INTERACTIONS_CSV.exists():
        return INTERACTIONS_CSV

    if KAGGLE_INPUT_DIR.exists():
        for candidate in KAGGLE_INPUT_DIR.glob("*.csv"):
            print(f"  Found CSV in Kaggle input: {candidate}")
            return candidate

    print(f"  CSV not found — downloading from Kaggle dataset '{KAGGLE_DATASET}' ...")
    try:
        import kaggle  # noqa: F401 — triggers credential check
        import subprocess
        dest = INTERACTIONS_CSV.parent
        dest.mkdir(parents=True, exist_ok=True)
        result = subprocess.run(
            ["kaggle", "datasets", "download", "-d", KAGGLE_DATASET, "--unzip", "-p", str(dest)],
            capture_output=True, text=True,
        )
        if result.returncode == 0:
            for f in dest.glob("*.csv"):
                print(f"  Downloaded: {f.name}")
                return f
        else:
            print(f"  Kaggle download failed: {result.stderr.strip()}")
    except ImportError:
        print("  kaggle package not installed — run: uv add kaggle")
    except Exception as e:
        print(f"  Could not download: {e}")

    return None


def import_drug_interactions(limit: int = None) -> int:
    """Import drug-drug interactions from the DDInter CSV (Kaggle dataset mghobashy/drug-drug-interactions)."""
    print("Importing drug interactions data...")

    filepath = _resolve_interactions_csv()
    if filepath is None:
        print("  Skipping drug interactions — CSV unavailable.")
        print(f"  To fix: attach the Kaggle dataset '{KAGGLE_DATASET}' to your notebook,")
        print("  or set up ~/.kaggle/kaggle.json for API access.")
        return 0

    total_records = 0
    for chunk in pd.read_csv(filepath, chunksize=10000):
        chunk.columns = [col.strip().lower().replace(' ', '_') for col in chunk.columns]

        records = []
        for _, row in chunk.iterrows():
            drug_1 = str(row.get('drug_1', row.get('drug1', row.iloc[0] if len(row) > 0 else '')))
            drug_2 = str(row.get('drug_2', row.get('drug2', row.iloc[1] if len(row) > 1 else '')))
            description = str(row.get('interaction_description', row.get('description',
                             row.get('interaction', row.iloc[2] if len(row) > 2 else ''))))
            if drug_1 and drug_2:
                records.append((drug_1, drug_2, description, classify_severity(description)))

        if records:
            execute_many(
                "INSERT INTO drug_interactions (drug_1, drug_2, interaction_description, severity) VALUES (?, ?, ?, ?)",
                records,
            )
            total_records += len(records)

        if limit and total_records >= limit:
            break

    print(f"  Imported {total_records} drug interaction records")
    return total_records


def import_all_data(interactions_limit: int = None) -> dict:
    """Initialize the database and import all structured data sources."""
    print(f"\n{'='*50}")
    print("AMR-Guard Data Import")
    print(f"{'='*50}\n")

    init_database()

    with get_connection() as conn:
        for table in ("eml_antibiotics", "atlas_susceptibility", "mic_breakpoints", "drug_interactions"):
            conn.execute(f"DELETE FROM {table}")
        conn.commit()
    print("Cleared existing data\n")

    results = {
        "eml_antibiotics": import_eml_antibiotics(),
        "atlas_susceptibility": import_atlas_susceptibility(),
        "mic_breakpoints": import_mic_breakpoints(),
        "drug_interactions": import_drug_interactions(limit=interactions_limit),
    }

    print(f"\n{'='*50}")
    print("Import Summary:")
    for table, count in results.items():
        print(f"  {table}: {count} records")
    print(f"{'='*50}\n")

    return results


if __name__ == "__main__":
    import_all_data(interactions_limit=50000)
